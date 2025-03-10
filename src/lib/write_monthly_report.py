import os
import openpyxl
import openpyxl.cell
from datetime import date, datetime
from typing import List
from .daily_work_info import DailyWorkInfo


TARGET_MONTH_CELL_POSITION = "J3"
PARTICIPATING_COMPANY_CELL_POSITION = "J4"
WORKER_CELL_POSITION = "J6"
DATE_COL = "B"
START_AT_COL = "E"
END_AT_COL = "F"
REST_TIME_COL = "G"
WORK_TIME_COL = "H"
WORK_DETAILS_COL = "I"
NOTES_CELL_POSITION = "B45"
MONTHLY_REPORT_TEMPLATE_PATH = "{data_path}/templates/monthly_work_report_template.xlsx"
REPORT_SHEET_NAME = "作業報告書"
MONTHLY_REPORT_OUTPUT_PATH = "{data_path}/outputs/作業報告書_{worker}_{target_month}.xlsx"


def validate_report_within_target_month(target_month: str):
    data_path = os.getenv("ROOT_DIR") + "/data"
    monthly_report_template_path = MONTHLY_REPORT_TEMPLATE_PATH.format(data_path=data_path)

    wb = openpyxl.load_workbook(monthly_report_template_path, data_only=True)
    ws = wb[REPORT_SHEET_NAME]

    target_month_cell = ws[TARGET_MONTH_CELL_POSITION]
    if target_month_cell.value != datetime.strptime(target_month, "%Y%m"):
        raise ValueError(f"作業報告書({monthly_report_template_path})の対象年月を更新してください。")


def write_monthly_report(worker: str, participating_company: str, target_month: str, infos: List[DailyWorkInfo]):
    data_path = os.getenv("ROOT_DIR") + "/data"
    monthly_report_template_path = MONTHLY_REPORT_TEMPLATE_PATH.format(data_path=data_path)

    wb = openpyxl.load_workbook(monthly_report_template_path, data_only=True)
    ws = wb[REPORT_SHEET_NAME]

    participating_company_cell = ws[PARTICIPATING_COMPANY_CELL_POSITION]
    participating_company_cell.value = participating_company

    worker_name_cell = ws[WORKER_CELL_POSITION]
    worker_name_cell.value = worker

    date_col = ws[DATE_COL]

    for info in infos:
        for date_cell in date_col:
            if isinstance(date_cell, openpyxl.cell.MergedCell) or not isinstance(date_cell.value, date):
                continue

            date_str = date_cell.value.strftime("%Y-%m-%d")
            if date_str == info.date:
                start_at_cell = ws[f"{START_AT_COL}{date_cell.row}"]
                start_at_cell.value = info.start_at

                end_at_cell = ws[f"{END_AT_COL}{date_cell.row}"]
                end_at_cell.value = info.end_at

                rest_time_cell = ws[f"{REST_TIME_COL}{date_cell.row}"]
                rest_time_cell.value = info.rest_time

                work_time_cell = ws[f"{WORK_TIME_COL}{date_cell.row}"]
                work_time_cell.value = info.work_time

                work_details_cell = ws[f"{WORK_DETAILS_COL}{date_cell.row}"]
                work_details_cell.value = info.work_details

    notes_list = sum([info.notes for info in infos], [])
    notes = "\n".join(notes_list)
    notes_col = ws[NOTES_CELL_POSITION]
    notes_col.value = notes

    output_path = MONTHLY_REPORT_OUTPUT_PATH.format(data_path=data_path, worker=worker, target_month=target_month)
    wb.save(output_path)
    print(f"作業報告書を出力しました: {output_path}")
